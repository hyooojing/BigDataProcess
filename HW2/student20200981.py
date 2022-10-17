#!/usr/bin/pyhton3

import openpyxl

wb = openpyxl.load_workbook("student.xlsx")
ws = wb['Sheet1']

list = []
row_id = 1

for row in ws:
   if row_id != 1:
      sum_v = ws.cell(row = row_id, column = 3).value * 0.3
      sum_v += ws.cell(row = row_id, column = 4).value * 0.35
      sum_v += ws.cell(row = row_id, column = 5).value * 0.34
      sum_v += ws.cell(row = row_id, column = 6).value
      ws.cell(row = row_id, column = 7).value = sum_v
      total = ws.cell(row = row_id, column = 7).value
      list.append([row_id, total])
   row_id += 1
   
list2 = sorted(list, key=lambda x:x[1], reverse=True)
student = len(list)

for i in range(a):
   ws.cell(row = list2[i][0], column = 8).value = "A0"
for i in range(aPlus):
   ws.cell(row = list2[i][0], column = 8).value = "A+"
for i in range(a, a + b):
   ws.cell(row = list2[i][0], column = 8).value = "B0"
for i in range(a, a + bPlus):
   ws.cell(row = list2[i][0], column = 8).value = "B+"
for i in range(a + b, a + b + c):
   ws.cell(row = list2[i][0], column = 8).value = "C0"
for i in range(a + b, a + b + cPlus):
   ws.cell(row = list2[i][0], column = 8).value = "C+"

wb.save("student.xlsx")
